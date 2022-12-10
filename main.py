import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit


class Vacancy:
    to_rub = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.sal_from = int(float(vacancy['salary_from']))
        self.sal_to = int(float(vacancy['salary_to']))
        self.sal_curr = vacancy['salary_currency']
        self.sal_average = self.to_rub[self.sal_curr] * (self.sal_from + self.sal_to) / 2
        self.name_area = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.name_file = file_name
        self.name_vac = vacancy_name

    @staticmethod
    def incr(dic, key, amount):
        if key in dic:
            dic[key] += amount
        else:
            dic[key] = amount

    @staticmethod
    def average(dic):
        new_dic = {}
        for key, values in dic.items():
            new_dic[key] = int(sum(values) / len(values))
        return new_dic

    def reader(self):
        with open(self.name_file, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    def statistic(self):
        sal = {}
        sal_vac_name = {}
        sal_city = {}
        vac_count = 0

        for vac_dic in self.reader():
            vacancy = Vacancy(vac_dic)
            self.incr(sal, vacancy.year, [vacancy.sal_average])
            if vacancy.name.find(self.name_vac) != -1:
                self.incr(sal_vac_name, vacancy.year, [vacancy.sal_average])
            self.incr(sal_city, vacancy.name_area, [vacancy.sal_average])
            vac_count += 1

        vac_number = dict([(key, len(value)) for key, value in sal.items()])
        vac_number_name = dict([(key, len(value)) for key, value in sal_vac_name.items()])

        if not sal_vac_name:
            sal_vac_name = dict([(key, [0]) for key, value in sal.items()])
            vac_number_name = dict([(key, 0) for key, value in vac_number.items()])

        stat = self.average(sal)
        stat2 = self.average(sal_vac_name)
        stat3 = self.average(sal_city)

        stat4 = {}
        for year, salaries in sal_city.items():
            stat4[year] = round(len(salaries) / vac_count, 4)
        stat4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in stat4.items()]))
        stat4.sort(key=lambda a: a[-1], reverse=True)
        stat5 = stat4.copy()
        stat4 = dict(stat4)
        stat3 = list(filter(lambda a: a[0] in list(stat4.keys()), [(key, value) for key, value in stat3.items()]))
        stat3.sort(key=lambda a: a[-1], reverse=True)
        stat3 = dict(stat3[:10])
        stat5 = dict(stat5[:10])

        return stat, vac_number, stat2, vac_number_name, stat3, stat5

    @staticmethod
    def print_stat(stats1, stats2, stats3, stats4, stats5, stats6):
        print('Динамика уровня зарплат по годам: {0}'.format(stats1))
        print('Динамика количества вакансий по годам: {0}'.format(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stats1, stats2, stats3, stats4, stats5, stats6 = dataset.statistic()
        dataset.print_stat(stats1, stats2, stats3, stats4, stats5, stats6)

        report = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)
        report.generate_excel()
        report.image()
        report.pdf()


class Report:
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        self.web = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def generate_excel(self):
        web1 = self.web.active
        web1.title = 'Статистика по годам'
        web1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            web1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        widths_column = []
        for row in data:
            for i, cell in enumerate(row):
                if len(widths_column) > i:
                    if len(cell) > widths_column[i]:
                        widths_column[i] = len(cell)
                else:
                    widths_column += [len(cell)]

        for i, column_width in enumerate(widths_column, 1):  # ,1 to start at 1
            web1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        web2 = self.web.create_sheet('Статистика по городам')
        for row in data:
            web2.append(row)

        widths_column = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(widths_column) > i:
                    if len(cell) > widths_column[i]:
                        widths_column[i] = len(cell)
                else:
                    widths_column += [len(cell)]

        for i, column_width in enumerate(widths_column, 1):  # ,1 to start at 1
            web2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            web1[col + '1'].font = font_bold
            web2[col + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            web2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                web2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.stats1[1] = 1
        for row, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                web1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def image(self):
        fig, ((x1, x2), (x3, x4)) = plt.subplots(nrows=2, ncols=2)

        line1 = x1.bar(np.array(list(self.stats1.keys())) - 0.4, self.stats1.values(), width=0.4)
        line2 = x1.bar(np.array(list(self.stats1)), self.stats3.values(), width=0.4)
        x1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        x1.grid(axis='y')
        x1.legend((line1[0], line2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        x1.set_xticks(np.array(list(self.stats1.keys())) - 0.2, list(self.stats1.keys()), rotation=90)
        x1.xaxis.set_tick_params(labelsize=8)
        x1.yaxis.set_tick_params(labelsize=8)

        x2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        line1 = x2.bar(np.array(list(self.stats2.keys())) - 0.4, self.stats2.values(), width=0.4)
        line2 = x2.bar(np.array(list(self.stats2)), self.stats4.values(), width=0.4)
        x2.legend((line1[0], line2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                   prop={'size': 8})
        x2.set_xticks(np.array(list(self.stats2.keys())) - 0.2, list(self.stats2.keys()), rotation=90)
        x2.grid(axis='y')
        x2.xaxis.set_tick_params(labelsize=8)
        x2.yaxis.set_tick_params(labelsize=8)

        x3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        x3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats5.keys()))]),
                 list(reversed(list(self.stats5.values()))), color='blue', height=0.5, align='center')
        x3.yaxis.set_tick_params(labelsize=6)
        x3.xaxis.set_tick_params(labelsize=8)
        x3.grid(axis='x')

        x4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.stats6.values()])
        x4.pie(list(self.stats6.values()) + [other], labels=list(self.stats6.keys()) + ['Другие'],
                textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def pdf(self):
        env = Environment(loader=FileSystemLoader('../templates'))
        template = env.get_template("pdf_template.html")
        stats = []
        for year in self.stats1.keys():
            stats.append([year, self.stats1[year], self.stats2[year], self.stats3[year], self.stats4[year]])
        for key in self.stats6:
            self.stats6[key] = round(self.stats6[key] * 100, 2)
        pdf_template = template.render(
            {'name': self.vacancy_name, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
             'stats': stats, 'stats5': self.stats5, 'stats6': self.stats6})
        config = pdfkit.configuration(wkhtmltopdf=r'/usr/bin/wkhtmltopdf')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})
        pdfkit.from_string(pdf_template, 'report.pdf', options={"enable-local-file-access": ""})


if __name__ == '__main__':
        InputConnect()



